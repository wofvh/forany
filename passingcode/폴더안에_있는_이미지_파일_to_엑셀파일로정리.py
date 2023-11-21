import os
import openpyxl

def create_excel_file():
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Image Names"
    excel_file_path = "D:/coco128/F1score/test/image_names.xlsx"
    # "safe" 폴더의 이미지 파일 목록 가져오기
    safe_folder_path = "D:\coco128/F1score/test\safe/"
    safe_image_names = os.listdir(safe_folder_path)

    # "unsafe" 폴더의 이미지 파일 목록 가져오기
    unsafe_folder_path = "D:\coco128/F1score/test/unsafe/"
    unsafe_image_names = os.listdir(unsafe_folder_path)

    # 각 이미지 파일 이름을 엑셀 파일에 추가
    for idx, image_name in enumerate(safe_image_names, start=1):
        ws.cell(row=idx, column=1, value=image_name)
        ws.cell(row=idx, column=2, value="safe")

    for idx, image_name in enumerate(unsafe_image_names, start=len(safe_image_names) + 2):
        ws.cell(row=idx, column=1, value=image_name)
        ws.cell(row=idx, column=2, value="unsafe")

    # 엑셀 파일 저장
    wb.save(excel_file_path)

if __name__ == "__main__":
    create_excel_file()
